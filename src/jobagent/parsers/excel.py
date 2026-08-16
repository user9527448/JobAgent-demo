"""Deterministic XLSX position-table parsing with cell-level evidence."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO
from typing import Any, Final
from zipfile import BadZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]

from jobagent.core.exceptions import JsonValue
from jobagent.parsers.contracts import (
    CellRangeLocation,
    ParseErrorCode,
    ParseIssue,
    ParseRequest,
    ParseResult,
    ParseStatus,
    TableBlock,
    TableCell,
)

XLSX_MEDIA_TYPE: Final = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXCEL_PARSER_NAME: Final = "xlsx_position_table"

_POSITION_HEADERS: Final = frozenset(
    {
        "jobtitle",
        "position",
        "positiontitle",
        "role",
        "岗位",
        "岗位名称",
        "招聘岗位",
        "招聘职位",
        "职位",
        "职位名称",
    }
)
_RECOGNIZED_HEADERS: Final = _POSITION_HEADERS | frozenset(
    {
        "department",
        "education",
        "headcount",
        "location",
        "major",
        "number",
        "qualification",
        "requirements",
        "序号",
        "人数",
        "专业",
        "专业要求",
        "任职要求",
        "单位",
        "学历",
        "学历要求",
        "工作地点",
        "招聘人数",
        "条件",
        "用人单位",
        "部门",
    }
)
_IGNORED_HEADER_CHARACTERS: Final = frozenset(
    {":", "\uff1a", "(", ")", "\uff08", "\uff09", "/", "\\", "_", "-"}
)


@dataclass(frozen=True, slots=True)
class ExcelTablePolicy:
    """Bound deterministic header recognition for common position workbooks."""

    max_header_rows: int = 20
    min_recognized_header_cells: int = 2

    def __post_init__(self) -> None:
        if self.max_header_rows <= 0:
            raise ValueError("Excel header scan limit must be positive.")
        if self.min_recognized_header_cells < 2:
            raise ValueError("Excel header recognition requires at least two cells.")


@dataclass(frozen=True, slots=True)
class _HeaderRegion:
    row_number: int
    start_column: int
    end_column: int
    recognized_cells: int


class ExcelPositionTableParser:
    """Convert recognized XLSX position sheets into traceable table blocks."""

    def __init__(self, policy: ExcelTablePolicy | None = None) -> None:
        self._policy = policy or ExcelTablePolicy()

    @property
    def name(self) -> str:
        """Return the stable parser implementation name."""
        return EXCEL_PARSER_NAME

    @property
    def supported_media_types(self) -> tuple[str, ...]:
        """Return XLSX only; legacy XLS requires a separate proven dependency."""
        return (XLSX_MEDIA_TYPE,)

    def parse(self, request: ParseRequest) -> ParseResult:
        """Parse every recognized worksheet or return safe review diagnostics."""
        if request.source.media_type != XLSX_MEDIA_TYPE:
            return _failed_result(
                request,
                code=ParseErrorCode.INVALID_INPUT,
                message="The Excel parser only accepts the XLSX media type.",
            )
        try:
            workbook = load_workbook(
                BytesIO(request.content),
                read_only=False,
                data_only=True,
            )
        except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError):
            return _failed_result(
                request,
                code=ParseErrorCode.CORRUPT_DOCUMENT,
                message="The workbook is corrupt, encrypted, or not a valid XLSX document.",
            )

        blocks: list[TableBlock] = []
        issues: list[ParseIssue] = []
        worksheet_metadata: list[JsonValue] = []
        try:
            for worksheet in workbook.worksheets:
                header = _detect_header(worksheet, self._policy)
                if header is None:
                    issues.append(_header_issue(worksheet.title, self._policy.max_header_rows))
                    worksheet_metadata.append(
                        {
                            "sheet_name": worksheet.title,
                            "status": "review_required",
                        }
                    )
                    continue
                block = _build_table_block(request, worksheet, header)
                if block is None:
                    issues.append(_header_issue(worksheet.title, self._policy.max_header_rows))
                    worksheet_metadata.append(
                        {
                            "sheet_name": worksheet.title,
                            "status": "review_required",
                            "header_row": header.row_number,
                            "reason": "no_data_rows",
                        }
                    )
                    continue
                blocks.append(block)
                worksheet_metadata.append(
                    {
                        "sheet_name": worksheet.title,
                        "status": "parsed",
                        "header_row": header.row_number,
                        "data_rows": len(block.rows) - 1,
                    }
                )
        except (KeyError, TypeError, ValueError):
            return _failed_result(
                request,
                code=ParseErrorCode.CORRUPT_DOCUMENT,
                message="The workbook structure or worksheet content is unreadable.",
            )
        finally:
            workbook.close()

        metadata: dict[str, JsonValue] = {
            "worksheet_count": len(worksheet_metadata),
            "parsed_worksheet_count": len(blocks),
            "worksheets": worksheet_metadata,
        }
        if not blocks:
            return ParseResult(
                source=request.source,
                status=ParseStatus.FAILED,
                parser_name=self.name,
                issues=tuple(issues),
                metadata=metadata,
            )
        return ParseResult(
            source=request.source,
            status=ParseStatus.PARSED,
            parser_name=self.name,
            blocks=tuple(blocks),
            issues=tuple(issues),
            metadata=metadata,
        )


def _detect_header(worksheet: Any, policy: ExcelTablePolicy) -> _HeaderRegion | None:
    candidates: list[_HeaderRegion] = []
    merged_ranges = tuple(worksheet.merged_cells.ranges)
    final_row = min(worksheet.max_row, policy.max_header_rows)
    for row_number in range(1, final_row + 1):
        values = [
            _normalize_cell_value(
                _resolved_cell_value(worksheet, merged_ranges, row_number, column)
            )
            for column in range(1, worksheet.max_column + 1)
        ]
        nonempty_columns = [index + 1 for index, value in enumerate(values) if value]
        if not nonempty_columns:
            continue
        header_keys = {_header_key(values[column - 1]) for column in nonempty_columns}
        recognized = len(header_keys & _RECOGNIZED_HEADERS)
        if recognized >= policy.min_recognized_header_cells and bool(
            header_keys & _POSITION_HEADERS
        ):
            candidates.append(
                _HeaderRegion(
                    row_number=row_number,
                    start_column=min(nonempty_columns),
                    end_column=max(nonempty_columns),
                    recognized_cells=recognized,
                )
            )
    if not candidates:
        return None
    return max(
        candidates, key=lambda candidate: (candidate.recognized_cells, -candidate.row_number)
    )


def _build_table_block(
    request: ParseRequest,
    worksheet: Any,
    header: _HeaderRegion,
) -> TableBlock | None:
    source_rows = [header.row_number]
    source_rows.extend(
        row_number
        for row_number in range(header.row_number + 1, worksheet.max_row + 1)
        if any(
            _normalize_cell_value(worksheet.cell(row=row_number, column=column).value)
            for column in range(header.start_column, header.end_column + 1)
        )
    )
    if len(source_rows) == 1:
        return None

    merged_ranges = tuple(worksheet.merged_cells.ranges)
    rows = tuple(
        tuple(
            _table_cell(request, worksheet, merged_ranges, row_number, column)
            for column in range(header.start_column, header.end_column + 1)
        )
        for row_number in source_rows
    )
    start_cell = f"{get_column_letter(header.start_column)}{header.row_number}"
    end_cell = f"{get_column_letter(header.end_column)}{source_rows[-1]}"
    source_rows_json: list[JsonValue] = list(source_rows)
    return TableBlock(
        rows=rows,
        location=CellRangeLocation(
            source=request.source,
            sheet_name=worksheet.title,
            start_cell=start_cell,
            end_cell=end_cell,
        ),
        metadata={
            "header_row": header.row_number,
            "data_start_row": source_rows[1],
            "data_end_row": source_rows[-1],
            "source_rows": source_rows_json,
            "blank_rows_skipped": source_rows[-1] - source_rows[0] + 1 - len(source_rows),
        },
    )


def _table_cell(
    request: ParseRequest,
    worksheet: Any,
    merged_ranges: tuple[Any, ...],
    row_number: int,
    column_number: int,
) -> TableCell:
    value = worksheet.cell(row=row_number, column=column_number).value
    start_cell = f"{get_column_letter(column_number)}{row_number}"
    end_cell = start_cell
    for merged_range in merged_ranges:
        min_column, min_row, max_column, max_row = merged_range.bounds
        if min_row <= row_number <= max_row and min_column <= column_number <= max_column:
            value = worksheet.cell(row=min_row, column=min_column).value
            start_cell = f"{get_column_letter(min_column)}{min_row}"
            end_cell = f"{get_column_letter(max_column)}{max_row}"
            break
    return TableCell(
        value=_normalize_cell_value(value),
        location=CellRangeLocation(
            source=request.source,
            sheet_name=worksheet.title,
            start_cell=start_cell,
            end_cell=end_cell,
        ),
    )


def _resolved_cell_value(
    worksheet: Any,
    merged_ranges: tuple[Any, ...],
    row_number: int,
    column_number: int,
) -> object:
    for merged_range in merged_ranges:
        min_column, min_row, max_column, max_row = merged_range.bounds
        if min_row <= row_number <= max_row and min_column <= column_number <= max_column:
            return worksheet.cell(row=min_row, column=min_column).value
    return worksheet.cell(row=row_number, column=column_number).value


def _normalize_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.split())
    if isinstance(value, datetime) and value.time() == time():
        return value.date().isoformat()
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return str(value).lower()
    return str(value)


def _header_key(value: str) -> str:
    return "".join(
        character.casefold()
        for character in value
        if not character.isspace() and character not in _IGNORED_HEADER_CHARACTERS
    )


def _header_issue(sheet_name: str, inspected_rows: int) -> ParseIssue:
    return ParseIssue(
        code=ParseErrorCode.HEADER_NOT_RECOGNIZED,
        message=(
            f"Worksheet '{sheet_name}' requires manual review because no position header "
            "was recognized."
        ),
        details={
            "sheet_name": sheet_name,
            "inspected_rows": inspected_rows,
            "review_required": True,
        },
    )


def _failed_result(
    request: ParseRequest,
    *,
    code: ParseErrorCode,
    message: str,
) -> ParseResult:
    return ParseResult(
        source=request.source,
        status=ParseStatus.FAILED,
        parser_name=EXCEL_PARSER_NAME,
        issues=(ParseIssue(code=code, message=message),),
    )
