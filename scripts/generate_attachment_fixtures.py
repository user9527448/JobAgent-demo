"""Regenerate the sanitized JAI-016 attachment fixtures and reviewed snapshots."""

from __future__ import annotations

import json
from collections.abc import Callable
from datetime import date
from io import BytesIO
from pathlib import Path

import pymupdf
from openpyxl import Workbook  # type: ignore[import-untyped]

from jobagent.parsers import (
    PDF_MEDIA_TYPE,
    XLSX_MEDIA_TYPE,
    ParseRequest,
    ParseSource,
    ParseSourceType,
    build_parser_registry,
    serialize_parse_result,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
FIXTURE_DIR = PROJECT_ROOT / "tests" / "fixtures" / "attachments"


def _pdf_bytes(pages: tuple[str, ...]) -> bytes:
    document = pymupdf.open()  # type: ignore[no-untyped-call]
    try:
        for text in pages:
            page = document.new_page(width=595, height=842)
            if text:
                page.insert_textbox(
                    pymupdf.Rect(72, 72, 523, 770),  # type: ignore[no-untyped-call]
                    text,
                    fontsize=11,
                )
        content: bytes = document.tobytes(  # type: ignore[no-untyped-call]
            garbage=4,
            deflate=True,
        )
        return content
    finally:
        document.close()  # type: ignore[no-untyped-call]


def _xlsx_bytes(configure: Callable[[Workbook], None]) -> bytes:
    workbook = Workbook()
    configure(workbook)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _xlsx_basic(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "岗位表"
    sheet.append(["岗位名称", "招聘人数", "工作地点", "截止日期"])
    sheet.append(["软件工程师", 2, "上海", date(2026, 9, 1)])


def _xlsx_english(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Roles"
    sheet.append(["Position", "Department", "Location", "Application Date"])
    sheet.append(["Analyst", "Data", "Beijing", "2026/09/08"])


def _xlsx_multiple_sheets(workbook: Workbook) -> None:
    first = workbook.active
    first.title = "技术"
    first.append(["岗位", "部门", "截止日期"])
    first.append(["开发", "研发", "2026年9月15日"])
    second = workbook.create_sheet("运营")
    second.append(["职位名称", "人数"])
    second.append(["运营专员", 1])


def _xlsx_merged_and_blank(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "计划"
    sheet.append(["岗位名称", "部门", "招聘人数"])
    sheet.append(["工程师", "平台", 1])
    sheet.append([None, "数据", 2])
    sheet.merge_cells("A2:A3")
    sheet.append([])
    sheet.append(["设计师", "产品", 1])


def _xlsx_review(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "说明"
    sheet.append(["模板说明", "备注"])
    sheet.append(["仅供离线测试", "无岗位表头"])


def main() -> int:
    """Write deterministic, synthetic fixtures and snapshot their intermediate results."""
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    fixtures = (
        (
            "pdf-basic",
            "pdf-basic.pdf",
            PDF_MEDIA_TYPE,
            _pdf_bytes(("Public recruitment notice. " * 8,)),
        ),
        (
            "pdf-multipage",
            "pdf-multipage.pdf",
            PDF_MEDIA_TYPE,
            _pdf_bytes(("First page position details. " * 6, "Second page requirements. " * 6)),
        ),
        (
            "pdf-line-normalization",
            "pdf-line-normalization.pdf",
            PDF_MEDIA_TYPE,
            _pdf_bytes(("Position A     Shanghai\nRequirements     public sample " * 5,)),
        ),
        ("pdf-sparse", "pdf-sparse.pdf", PDF_MEDIA_TYPE, _pdf_bytes(("Role",))),
        ("pdf-blank", "pdf-blank.pdf", PDF_MEDIA_TYPE, _pdf_bytes(("",))),
        ("xlsx-basic-cn", "xlsx-basic-cn.xlsx", XLSX_MEDIA_TYPE, _xlsx_bytes(_xlsx_basic)),
        (
            "xlsx-english",
            "xlsx-english.xlsx",
            XLSX_MEDIA_TYPE,
            _xlsx_bytes(_xlsx_english),
        ),
        (
            "xlsx-multiple-sheets",
            "xlsx-multiple-sheets.xlsx",
            XLSX_MEDIA_TYPE,
            _xlsx_bytes(_xlsx_multiple_sheets),
        ),
        (
            "xlsx-merged-blank",
            "xlsx-merged-blank.xlsx",
            XLSX_MEDIA_TYPE,
            _xlsx_bytes(_xlsx_merged_and_blank),
        ),
        ("xlsx-review", "xlsx-review.xlsx", XLSX_MEDIA_TYPE, _xlsx_bytes(_xlsx_review)),
    )
    registry = build_parser_registry()
    cases: list[dict[str, object]] = []
    for source_id, (case_id, filename, media_type, content) in enumerate(fixtures, start=1):
        (FIXTURE_DIR / filename).write_bytes(content)
        source = ParseSource(
            source_type=ParseSourceType.ATTACHMENT,
            source_id=source_id,
            source_name=filename,
            media_type=media_type,
        )
        result = registry.parse(ParseRequest(source=source, content=content))
        cases.append(
            {
                "id": case_id,
                "file": filename,
                "media_type": media_type,
                "expected": serialize_parse_result(result),
            }
        )
    manifest = {"schema_version": 1, "cases": cases}
    (FIXTURE_DIR / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
