from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]

from jobagent.parsers import (
    EXCEL_PARSER_NAME,
    XLSX_MEDIA_TYPE,
    CellRangeLocation,
    ExcelPositionTableParser,
    ExcelTablePolicy,
    ParseErrorCode,
    ParseRequest,
    ParseSource,
    ParseSourceType,
    ParseStatus,
    TableBlock,
    build_parser_registry,
)


def _source(*, media_type: str = XLSX_MEDIA_TYPE) -> ParseSource:
    return ParseSource(
        source_type=ParseSourceType.ATTACHMENT,
        source_id=15,
        source_name="objects/positions.xlsx",
        media_type=media_type,
    )


def _workbook_bytes(configure: object) -> bytes:
    workbook = Workbook()
    callback = configure
    assert callable(callback)
    callback(workbook)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def test_parser_extracts_multiple_sheets_blank_rows_and_merged_cell_evidence() -> None:
    def configure(workbook: object) -> None:
        positions = workbook.active  # type: ignore[attr-defined]
        positions.title = "岗位表"
        positions.merge_cells("A1:D1")
        positions["A1"] = "2026 年招聘计划"
        positions.append([])
        positions.append(["岗位名称", "部门", "招聘人数", "工作地点"])
        positions.append(["工程师", "研发", 2, "上海"])
        positions.append([None, "平台", 1, "北京"])
        positions.merge_cells("A4:A5")
        positions.append([])
        positions.append(["分析师", "数据", 1, date(2026, 8, 16)])

        english = workbook.create_sheet("Roles")  # type: ignore[attr-defined]
        english.append(["Position", "Location"])
        english.append(["Designer", "Shenzhen"])

        instructions = workbook.create_sheet("说明")  # type: ignore[attr-defined]
        instructions.append(["请勿修改模板说明"])

    content = _workbook_bytes(configure)
    result = ExcelPositionTableParser().parse(ParseRequest(source=_source(), content=content))

    assert result.status is ParseStatus.PARSED
    assert result.parser_name == EXCEL_PARSER_NAME
    assert len(result.blocks) == 2
    assert len(result.issues) == 1
    assert result.issues[0].code is ParseErrorCode.HEADER_NOT_RECOGNIZED
    assert result.issues[0].details["sheet_name"] == "说明"
    assert result.metadata["worksheet_count"] == 3
    assert result.metadata["parsed_worksheet_count"] == 2

    table = result.blocks[0]
    assert isinstance(table, TableBlock)
    assert [[cell.value for cell in row] for row in table.rows] == [
        ["岗位名称", "部门", "招聘人数", "工作地点"],
        ["工程师", "研发", "2", "上海"],
        ["工程师", "平台", "1", "北京"],
        ["分析师", "数据", "1", "2026-08-16"],
    ]
    assert table.metadata["header_row"] == 3
    assert table.metadata["source_rows"] == [3, 4, 5, 7]
    assert table.metadata["blank_rows_skipped"] == 1
    assert isinstance(table.location, CellRangeLocation)
    assert (table.location.sheet_name, table.location.start_cell, table.location.end_cell) == (
        "岗位表",
        "A3",
        "D7",
    )
    merged_location = table.rows[2][0].location
    assert isinstance(merged_location, CellRangeLocation)
    assert (merged_location.start_cell, merged_location.end_cell) == ("A4", "A5")
    exact_location = table.rows[3][1].location
    assert isinstance(exact_location, CellRangeLocation)
    assert (exact_location.start_cell, exact_location.end_cell) == ("B7", "B7")


def test_parser_marks_unrecognized_headers_for_manual_review() -> None:
    def configure(workbook: object) -> None:
        sheet = workbook.active  # type: ignore[attr-defined]
        sheet.title = "Sheet1"
        sheet.append(["说明", "备注"])
        sheet.append(["模板", "待填写"])

    content = _workbook_bytes(configure)
    result = ExcelPositionTableParser().parse(ParseRequest(source=_source(), content=content))

    assert result.status is ParseStatus.FAILED
    assert result.blocks == ()
    assert result.issues[0].code is ParseErrorCode.HEADER_NOT_RECOGNIZED
    assert result.issues[0].details == {
        "sheet_name": "Sheet1",
        "inspected_rows": 20,
        "review_required": True,
    }
    assert result.metadata["parsed_worksheet_count"] == 0


def test_parser_recognizes_a_two_level_header_with_vertical_merges() -> None:
    def configure(workbook: object) -> None:
        sheet = workbook.active  # type: ignore[attr-defined]
        sheet.merge_cells("A1:A2")
        sheet["A1"] = "岗位名称"
        sheet["B1"] = "岗位信息"
        sheet["B2"] = "部门"
        sheet["C2"] = "招聘人数"
        sheet.append(["工程师", "研发", 2])

    content = _workbook_bytes(configure)
    result = ExcelPositionTableParser().parse(ParseRequest(source=_source(), content=content))

    assert result.status is ParseStatus.PARSED
    table = result.blocks[0]
    assert isinstance(table, TableBlock)
    assert table.metadata["header_row"] == 2
    assert [cell.value for cell in table.rows[0]] == ["岗位名称", "部门", "招聘人数"]


def test_parser_marks_recognized_header_without_data_for_review() -> None:
    def configure(workbook: object) -> None:
        sheet = workbook.active  # type: ignore[attr-defined]
        sheet.append(["岗位名称", "招聘人数"])

    content = _workbook_bytes(configure)
    result = ExcelPositionTableParser().parse(ParseRequest(source=_source(), content=content))

    assert result.status is ParseStatus.FAILED
    assert result.issues[0].code is ParseErrorCode.HEADER_NOT_RECOGNIZED
    worksheets = result.metadata["worksheets"]
    assert isinstance(worksheets, list)
    worksheet = worksheets[0]
    assert isinstance(worksheet, dict)
    assert worksheet["reason"] == "no_data_rows"


def test_parser_returns_safe_diagnostics_for_invalid_inputs() -> None:
    corrupt = ExcelPositionTableParser().parse(
        ParseRequest(source=_source(), content=b"not an xlsx archive")
    )
    wrong_media = ExcelPositionTableParser().parse(
        ParseRequest(
            source=_source(media_type="application/vnd.ms-excel"),
            content=b"legacy xls",
        )
    )

    assert corrupt.status is ParseStatus.FAILED
    assert corrupt.issues[0].code is ParseErrorCode.CORRUPT_DOCUMENT
    assert wrong_media.status is ParseStatus.FAILED
    assert wrong_media.issues[0].code is ParseErrorCode.INVALID_INPUT


def test_production_registry_registers_xlsx_but_not_legacy_xls() -> None:
    registry = build_parser_registry()

    assert registry.select(XLSX_MEDIA_TYPE) is not None
    result = registry.parse(
        ParseRequest(
            source=_source(media_type="application/vnd.ms-excel"),
            content=b"legacy xls",
        )
    )

    assert result.status is ParseStatus.UNSUPPORTED
    assert result.issues[0].code is ParseErrorCode.UNSUPPORTED_MEDIA_TYPE


@pytest.mark.parametrize(
    "policy",
    [
        lambda: ExcelTablePolicy(max_header_rows=0),
        lambda: ExcelTablePolicy(min_recognized_header_cells=1),
    ],
)
def test_policy_rejects_invalid_thresholds(policy: object) -> None:
    factory = policy
    assert callable(factory)
    with pytest.raises(ValueError):
        factory()
